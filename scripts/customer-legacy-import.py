#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json, execute_values


COMBINED_WS_HEADERS = [
    "No.",
    "Description",
    "Make",
    "Vehicle Year",
    "Serial No.",
    "Location Code",
    "On Rent",
    "FA Class Code",
    "FA Subclass Code",
    "Vehicle Registration No.",
    "Acquired",
    "Disposed",
    "Next Service Date",
    "Year",
    "Make2",
    "Purchase Date",
    "Vendor",
    "Purchase Price",
    "Customer #",
    "Corporate Name",
    "Location",
    "Year Category",
    "FA Class Name",
    "FA Subclass Name",
    "Rental $",
]

CUSTOMER_UNIT_HEADERS = [
    "NoToPrint",
    "FORMAT_TODAY_0_4_",
    "Customer__No__",
    "Rental_Header_No_",
    "Rental_Header_Sell_to_Customer_No_",
    "DescriptionToPrint",
]

INTERNAL_CUSTOMER_CODES = {"INTERNAL", "METRO"}
EXCLUDED_MARKET_LABELS = {"", "BANK", "LOST", "SALE", "SOLD"}
EXCLUDED_BRANCH_CODES = {"BANK", "LOST", "SALE", "SOLD"}
SOURCE = "legacy-customer-import"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Profile and seed customers from the legacy Fixed Asset Units workbook."
        )
    )
    parser.add_argument(
        "--source",
        required=True,
        help="Path to the Fixed Asset Units with Attribute Values workbook.",
    )
    parser.add_argument(
        "--out",
        help="Optional output directory for the exploration and seed report.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Write supported customer data into the database. Omit for dry-run.",
    )
    return parser.parse_args()


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if value.startswith(("'", '"')) and value.endswith(("'", '"')):
            value = value[1:-1]
        os.environ.setdefault(key, value)


def require_database_url(project_root: Path) -> str:
    load_dotenv(project_root / ".env")
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is required. Set it in the environment or .env.")
    return normalize_database_url(database_url)


def normalize_database_url(database_url: str) -> str:
    parsed = urlsplit(database_url)
    query_items = [(key, value) for key, value in parse_qsl(parsed.query) if key != "schema"]
    if not query_items:
        return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", parsed.fragment))
    return urlunsplit(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            urlencode(query_items),
            parsed.fragment,
        )
    )


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text == "None" else text


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def canonical_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def expected_header_map(headers: list[str], expected: list[str], sheet_name: str) -> dict[str, int]:
    if headers != expected:
        raise RuntimeError(
            f'Unexpected headers in "{sheet_name}". Expected {expected}, received {headers}.'
        )
    return {header: position for position, header in enumerate(headers)}


def choose_primary(counter: Counter[str]) -> str:
    if not counter:
        return ""
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))[0][0]


def infer_customer_type(customer_number: str, customer_name: str) -> str:
    if customer_number in INTERNAL_CUSTOMER_CODES:
        return "internal"
    if "metro trailer" in customer_name.lower():
        return "internal"
    return "commercial"


def normalize_market_label(value: str) -> str:
    return compact_spaces(value.replace("\n", " "))


def placeholder_address(city: str, state: str | None = None) -> dict[str, str]:
    return {
        "line1": "Legacy import placeholder",
        "city": city or "Unknown",
        "state": state or "NA",
        "postalCode": "00000",
        "country": "US",
    }


def ensure_output_directory(project_root: Path, output_override: str | None) -> Path:
    if output_override:
        output_dir = Path(output_override)
    else:
        timestamp = now_utc().strftime("%Y-%m-%dT%H-%M-%SZ")
        output_dir = project_root / "artifacts" / "customer-legacy-import" / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def parse_locations_sheet(workbook: Any) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    if "Locations" not in workbook.sheetnames:
        return {}, {}

    worksheet = workbook["Locations"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}, {}

    yard_code_index = branch_code_index = name_index = location_index = None
    header_row_number = None
    for position, row in enumerate(rows):
        normalized = [normalize_text(value) for value in row]
        if "Yard Code" in normalized and "Branch Code" in normalized:
            yard_code_index = normalized.index("Yard Code")
            branch_code_index = normalized.index("Branch Code")
            name_index = normalized.index("Name") if "Name" in normalized else None
            location_index = normalized.index("Location") if "Location" in normalized else None
            header_row_number = position
            break

    if header_row_number is None or yard_code_index is None or branch_code_index is None:
        return {}, {}

    yard_map: dict[str, dict[str, str]] = {}
    location_to_branch: dict[str, str] = {}
    for row in rows[header_row_number + 1 :]:
        yard_code = normalize_text(row[yard_code_index])
        branch_code = normalize_text(row[branch_code_index])
        if not yard_code or not branch_code:
            continue

        location_name = normalize_market_label(
            normalize_text(row[location_index]) if location_index is not None else ""
        )
        yard_name = compact_spaces(
            normalize_text(row[name_index]) if name_index is not None else ""
        )

        yard_map[yard_code] = {
            "branchCode": branch_code,
            "yardName": yard_name,
            "marketLabel": location_name,
        }
        if location_name:
            location_to_branch.setdefault(canonical_key(location_name), branch_code)

    return yard_map, location_to_branch


def parse_customer_units_sheet(workbook: Any) -> dict[str, set[str]]:
    if "Customer Unit #" not in workbook.sheetnames:
        return {}

    worksheet = workbook["Customer Unit #"]
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return {}

    headers = [normalize_text(value) for value in header_row]
    header_index = expected_header_map(headers, CUSTOMER_UNIT_HEADERS, "Customer Unit #")

    contract_numbers_by_customer: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        customer_number = normalize_text(row[header_index["Customer__No__"]])
        contract_number = normalize_text(row[header_index["Rental_Header_No_"]])
        if not customer_number or not contract_number:
            continue
        contract_numbers_by_customer[customer_number].add(contract_number)

    return contract_numbers_by_customer


def parse_combined_ws(
    workbook: Any,
    yard_map: dict[str, dict[str, str]],
    location_to_branch: dict[str, str],
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    if "Combined WS" not in workbook.sheetnames:
        raise RuntimeError('Workbook does not contain required sheet "Combined WS".')

    worksheet = workbook["Combined WS"]
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise RuntimeError('"Combined WS" is empty.')

    headers = [normalize_text(value) for value in header_row]
    header_index = expected_header_map(headers, COMBINED_WS_HEADERS, "Combined WS")

    customer_rollups: dict[str, dict[str, Any]] = {}
    unsupported_counts: Counter[str] = Counter()
    source_branch_counts: Counter[str] = Counter()
    source_market_counts: Counter[str] = Counter()
    row_count = 0

    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue

        row_count += 1
        customer_number = normalize_text(row[header_index["Customer #"]])
        customer_name = compact_spaces(normalize_text(row[header_index["Corporate Name"]]))
        location_code = normalize_text(row[header_index["Location Code"]])
        market_label = normalize_market_label(normalize_text(row[header_index["Location"]]))
        asset_number = normalize_text(row[header_index["No."]])

        if not customer_number or not customer_name:
            unsupported_counts["missing customer number or customer name"] += 1
            continue
        if customer_number == "Not On Rent" or customer_name == "Not On Rent":
            unsupported_counts["not on rent placeholder"] += 1
            continue

        customer_entry = customer_rollups.setdefault(
            customer_number,
            {
                "customerNumber": customer_number,
                "nameCounts": Counter(),
                "marketCounts": Counter(),
                "branchCounts": Counter(),
                "locationCodes": Counter(),
                "assetNumbers": set(),
                "locationSeeds": {},
                "sampleRows": [],
            },
        )

        customer_entry["nameCounts"][customer_name] += 1
        if market_label and market_label.upper() not in EXCLUDED_MARKET_LABELS:
            customer_entry["marketCounts"][market_label] += 1
            source_market_counts[market_label] += 1
        if location_code:
            customer_entry["locationCodes"][location_code] += 1

        mapped_branch_code = ""
        if location_code and location_code in yard_map:
            mapped_branch_code = yard_map[location_code]["branchCode"]
        elif market_label:
            mapped_branch_code = location_to_branch.get(canonical_key(market_label), "")
        if mapped_branch_code:
            customer_entry["branchCounts"][mapped_branch_code] += 1
            source_branch_counts[mapped_branch_code] += 1

        if asset_number:
            customer_entry["assetNumbers"].add(asset_number)

        if market_label and market_label.upper() not in EXCLUDED_MARKET_LABELS:
            location_seed = customer_entry["locationSeeds"].setdefault(
                market_label,
                {
                    "name": market_label,
                    "branchCodeCounts": Counter(),
                    "assetCount": 0,
                    "locationCodeCounts": Counter(),
                },
            )
            if mapped_branch_code and mapped_branch_code not in EXCLUDED_BRANCH_CODES:
                location_seed["branchCodeCounts"][mapped_branch_code] += 1
            if location_code:
                location_seed["locationCodeCounts"][location_code] += 1
            location_seed["assetCount"] += 1

        if len(customer_entry["sampleRows"]) < 3:
            customer_entry["sampleRows"].append(
                {
                    "assetNumber": asset_number,
                    "locationCode": location_code,
                    "marketLabel": market_label,
                    "customerName": customer_name,
                }
            )

    report = {
        "sheet": "Combined WS",
        "rowCount": row_count,
        "headers": headers,
        "unsupportedReasonCounts": dict(unsupported_counts),
        "topBranchCoverage": source_branch_counts.most_common(20),
        "topMarketLabels": source_market_counts.most_common(20),
    }
    return customer_rollups, report


def build_customer_payloads(
    customer_rollups: dict[str, dict[str, Any]],
    contract_numbers_by_customer: dict[str, set[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    customers: list[dict[str, Any]] = []
    locations: list[dict[str, Any]] = []

    for customer_number, rollup in sorted(customer_rollups.items()):
        canonical_name = choose_primary(rollup["nameCounts"])
        dominant_market = choose_primary(rollup["marketCounts"])
        filtered_branch_counts = Counter(
            {
                code: count
                for code, count in rollup["branchCounts"].items()
                if code not in EXCLUDED_BRANCH_CODES
            }
        )
        dominant_branch = choose_primary(filtered_branch_counts)
        branch_coverage = sorted(
            code
            for code in filtered_branch_counts.keys()
            if code not in EXCLUDED_BRANCH_CODES
        )
        customer_type = infer_customer_type(customer_number, canonical_name)
        linked_contracts = sorted(contract_numbers_by_customer.get(customer_number, set()))

        notes_lines = [
            "Seeded from Fixed Asset Units with Attribute Values workbook.",
            "Billing address is a placeholder derived from the dominant legacy market label.",
            "Contracts were not seeded because the workbook does not contain reliable contract dates, billing cadence, payment terms, or line pricing.",
        ]
        if linked_contracts:
            notes_lines.append(
                f'Legacy rental headers observed: {", ".join(linked_contracts[:10])}'
                + (" ..." if len(linked_contracts) > 10 else "")
            )

        customer_payload = {
            "id": stable_id("customer_legacy", customer_number),
            "customerNumber": customer_number,
            "name": canonical_name,
            "customerType": customer_type,
            "contactInfo": {
                "name": canonical_name,
                "source": SOURCE,
                "placeholder": True,
            },
            "billingAddress": placeholder_address(dominant_market or "Unknown"),
            "branchCoverage": branch_coverage,
            "notes": " ".join(notes_lines),
            "assetCount": len(rollup["assetNumbers"]),
            "legacyMarketCount": len(rollup["locationSeeds"]),
            "legacyContractHeaders": linked_contracts,
            "sampleRows": rollup["sampleRows"],
            "dominantBranchCode": dominant_branch or None,
        }
        customers.append(customer_payload)

        for market_label, location_rollup in sorted(rollup["locationSeeds"].items()):
            dominant_location_branch = choose_primary(location_rollup["branchCodeCounts"])
            location_id = stable_id(
                "customer_location_legacy", f"{customer_number}:{market_label}"
            )
            locations.append(
                {
                    "id": location_id,
                    "customerId": customer_payload["id"],
                    "customerNumber": customer_number,
                    "name": f"Legacy market - {market_label}",
                    "address": placeholder_address(market_label),
                    "contactPerson": {
                        "name": canonical_name,
                        "source": SOURCE,
                        "placeholder": True,
                    },
                    "deliveryNotes": (
                        "Derived from legacy workbook market label. "
                        "This is not a verified delivery site address."
                    ),
                    "isPrimary": market_label == dominant_market,
                    "branchCode": dominant_location_branch or None,
                    "assetCount": location_rollup["assetCount"],
                    "sourceLocationCodes": sorted(location_rollup["locationCodeCounts"].keys()),
                }
            )

    return customers, locations


def write_report(
    output_dir: Path,
    report: dict[str, Any],
    customers: list[dict[str, Any]],
    locations: list[dict[str, Any]],
    write_mode: bool,
    seed_result: dict[str, Any] | None = None,
) -> None:
    summary_payload = {
        "writeMode": write_mode,
        "report": report,
        "seedResult": seed_result,
    }
    (output_dir / "summary.json").write_text(
        json.dumps(summary_payload, indent=2),
        encoding="utf-8",
    )
    (output_dir / "customers-preview.json").write_text(
        json.dumps(customers[:100], indent=2),
        encoding="utf-8",
    )
    (output_dir / "locations-preview.json").write_text(
        json.dumps(locations[:100], indent=2),
        encoding="utf-8",
    )


def build_report(
    source_path: Path,
    combined_report: dict[str, Any],
    yard_map: dict[str, dict[str, str]],
    contract_numbers_by_customer: dict[str, set[str]],
    customers: list[dict[str, Any]],
    locations: list[dict[str, Any]],
) -> dict[str, Any]:
    branch_coverage_counts: Counter[str] = Counter()
    customer_type_counts: Counter[str] = Counter()
    location_branch_counts: Counter[str] = Counter()

    for customer in customers:
        customer_type_counts[customer["customerType"]] += 1
        for branch_code in customer["branchCoverage"]:
            branch_coverage_counts[branch_code] += 1

    for location in locations:
        if location["branchCode"]:
            location_branch_counts[str(location["branchCode"])] += 1

    return {
        "generatedAt": now_utc().isoformat(),
        "source": str(source_path),
        "coverage": {
            "seedableCustomers": len(customers),
            "seedableCustomerLocations": len(locations),
            "seedableContractHeadersReferenceOnly": sum(
                len(contract_numbers) for contract_numbers in contract_numbers_by_customer.values()
            ),
            "supportedFirstClassFields": [
                "customers.customerNumber <- Combined WS Customer #",
                "customers.name <- dominant Combined WS Corporate Name",
                "customers.customerType <- inferred from customer number / name",
                "customers.billingAddress <- placeholder derived from dominant legacy market label",
                "customers.branchCoverage <- branches inferred from Locations sheet and Combined WS",
                "customers.notes <- explicit seed provenance and contract gap note",
                "customer_locations.name <- legacy market labels",
                "customer_locations.address <- placeholder derived from market label",
                "customer_locations.deliveryNotes <- explicit legacy placeholder notice",
            ],
            "notSeededAsFirstClassFields": [
                "contracts",
                "contract lines",
                "invoices",
                "billing cadence",
                "payment terms",
                "contract start dates",
                "contract end dates",
                "line pricing",
                "verified site addresses",
            ],
        },
        "locationsSheet": {
            "mappedYardCodes": len(yard_map),
        },
        "combinedSheet": combined_report,
        "customers": {
            "customerTypeCounts": dict(customer_type_counts),
            "topBranchCoverage": branch_coverage_counts.most_common(20),
            "topLocationBranchAssignments": location_branch_counts.most_common(20),
            "topCustomersByAssetCount": [
                [customer["customerNumber"], customer["name"], customer["assetCount"]]
                for customer in sorted(
                    customers,
                    key=lambda item: (-int(item["assetCount"]), item["customerNumber"]),
                )[:20]
            ],
        },
    }


def upsert_customers(cursor: Any, customers: list[dict[str, Any]], timestamp: datetime) -> None:
    customer_rows = [
        (
            customer["id"],
            customer["customerNumber"],
            customer["name"],
            customer["customerType"],
            Json(customer["contactInfo"]),
            Json(customer["billingAddress"]),
            Json(customer["branchCoverage"]),
            customer["notes"],
            timestamp,
            timestamp,
        )
        for customer in customers
    ]

    execute_values(
        cursor,
        """
        INSERT INTO customers (
          id,
          customer_number,
          name,
          customer_type,
          contact_info,
          billing_address,
          branch_coverage,
          notes,
          created_at,
          updated_at
        ) VALUES %s
        ON CONFLICT (customer_number) DO UPDATE SET
          name = EXCLUDED.name,
          customer_type = EXCLUDED.customer_type,
          contact_info = EXCLUDED.contact_info,
          billing_address = EXCLUDED.billing_address,
          branch_coverage = EXCLUDED.branch_coverage,
          notes = EXCLUDED.notes,
          updated_at = EXCLUDED.updated_at
        """,
        customer_rows,
        page_size=250,
    )


def fetch_customer_ids(cursor: Any, customer_numbers: list[str]) -> dict[str, str]:
    cursor.execute(
        """
        SELECT customer_number, id
        FROM customers
        WHERE customer_number = ANY(%s)
        """,
        (customer_numbers,),
    )
    return {customer_number: customer_id for customer_number, customer_id in cursor.fetchall()}


def upsert_customer_locations(
    cursor: Any,
    locations: list[dict[str, Any]],
    customer_id_by_number: dict[str, str],
    timestamp: datetime,
) -> None:
    location_rows = []
    for location in locations:
        customer_id = customer_id_by_number[location["customerNumber"]]
        location_rows.append(
            (
                location["id"],
                customer_id,
                location["name"],
                Json(location["address"]),
                Json(location["contactPerson"]),
                location["deliveryNotes"],
                location["isPrimary"],
                timestamp,
                timestamp,
            )
        )

    execute_values(
        cursor,
        """
        INSERT INTO customer_locations (
          id,
          customer_id,
          name,
          address,
          contact_person,
          delivery_notes,
          is_primary,
          created_at,
          updated_at
        ) VALUES %s
        ON CONFLICT (id) DO UPDATE SET
          customer_id = EXCLUDED.customer_id,
          name = EXCLUDED.name,
          address = EXCLUDED.address,
          contact_person = EXCLUDED.contact_person,
          delivery_notes = EXCLUDED.delivery_notes,
          is_primary = EXCLUDED.is_primary,
          updated_at = EXCLUDED.updated_at
        """,
        location_rows,
        page_size=500,
    )


def seed_database(
    database_url: str,
    customers: list[dict[str, Any]],
    locations: list[dict[str, Any]],
) -> dict[str, Any]:
    timestamp = now_utc()
    connection = psycopg2.connect(database_url)
    try:
        with connection:
            with connection.cursor() as cursor:
                upsert_customers(cursor, customers, timestamp)
                customer_id_by_number = fetch_customer_ids(
                    cursor, [customer["customerNumber"] for customer in customers]
                )
                upsert_customer_locations(cursor, locations, customer_id_by_number, timestamp)

                cursor.execute("SELECT COUNT(*) FROM customers")
                customer_count = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM customer_locations")
                customer_location_count = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM contracts")
                contract_count = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM contract_lines")
                contract_line_count = cursor.fetchone()[0]

        return {
            "seededCustomers": len(customers),
            "seededCustomerLocations": len(locations),
            "databaseCustomerCount": customer_count,
            "databaseCustomerLocationCount": customer_location_count,
            "databaseContractCount": contract_count,
            "databaseContractLineCount": contract_line_count,
        }
    finally:
        connection.close()


def main() -> int:
    args = parse_args()
    project_root = Path.cwd()
    source_path = Path(args.source)
    if not source_path.exists():
        raise RuntimeError(f'Source workbook "{source_path}" does not exist.')

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    yard_map, location_to_branch = parse_locations_sheet(workbook)
    contract_numbers_by_customer = parse_customer_units_sheet(workbook)
    customer_rollups, combined_report = parse_combined_ws(
        workbook, yard_map, location_to_branch
    )
    customers, locations = build_customer_payloads(
        customer_rollups, contract_numbers_by_customer
    )

    report = build_report(
        source_path,
        combined_report,
        yard_map,
        contract_numbers_by_customer,
        customers,
        locations,
    )
    output_dir = ensure_output_directory(project_root, args.out)

    seed_result = None
    if args.write:
        database_url = require_database_url(project_root)
        seed_result = seed_database(database_url, customers, locations)

    write_report(output_dir, report, customers, locations, args.write, seed_result)

    payload = {
        "outputDirectory": str(output_dir),
        "writeMode": args.write,
        "seedableCustomers": report["coverage"]["seedableCustomers"],
        "seedableCustomerLocations": report["coverage"]["seedableCustomerLocations"],
        "seedableContractHeadersReferenceOnly": report["coverage"][
            "seedableContractHeadersReferenceOnly"
        ],
        "seedResult": seed_result,
    }
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover
        print(str(error), file=sys.stderr)
        raise
