#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json, execute_values


EXPECTED_HEADERS = [
    "No.",
    "Description",
    "Make",
    "Vehicle Year",
    "Serial No.",
    "MTRZ",
    "Location Code",
    "On Rent",
    "FA Class Code",
    "FA Subclass Code",
    "Vehicle Registration No.",
    "Acquired",
    "Disposed",
    "Next Service Date",
]

RETIRED_LOCATION_CODES = {"BANK", "LOST", "SALE", "SOLD"}
PHYSICAL_BRANCH_CODES = {
    "ALA",
    "ATL",
    "BHM",
    "CLE",
    "CLT",
    "DAL",
    "DFW",
    "GSP",
    "LSV",
    "MEM",
    "NVL",
    "ONT",
    "TPA",
    "TUP",
}
SPECIAL_LOCATION_CODES = {
    "AV CA",
    "BANK",
    "BC YARD",
    "CAL",
    "CF",
    "CLT-2",
    "DAMCO",
    "DFWJ",
    "ITS",
    "LOST",
    "SAG",
    "SALE",
    "SEV",
    "SMITH",
    "SOLD",
    "UPS",
}
SOURCE = "fixed-assets-import"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Profile and seed Metro Trailer assets from the legacy fixed-assets "
            "Excel export."
        ),
    )
    parser.add_argument(
        "--source",
        required=True,
        help="Path to the Fixed Assets .xlsx workbook.",
    )
    parser.add_argument(
        "--out",
        help="Optional output directory for the exploration and seed report.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Write the supported rows into the database. Omit for dry-run.",
    )
    return parser.parse_args()


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if value.startswith(("'", '"')) and value.endswith(("'", '"')):
            value = value[1:-1]
        os.environ.setdefault(key, value)


def require_database_url(project_root: Path) -> str:
    load_dotenv(project_root / ".env")
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is required. Set it in the environment or .env.")
    return normalize_database_url(database_url)


def normalize_database_url(database_url: str) -> str:
    parsed = urlsplit(database_url)
    query_items = [(key, value) for key, value in parse_qsl(parsed.query) if key != "schema"]
    if not query_items:
        return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", parsed.fragment))
    return urlunsplit(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            urlencode(query_items),
            parsed.fragment,
        )
    )


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-") or "unknown"


def stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text == "None" else text


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return normalize_text(value).lower() == "true"


def normalize_year(value: Any) -> int | None:
    text = normalize_text(value)
    if not text or text in {"0", "0.0"}:
        return None
    try:
        year = int(float(text))
    except ValueError:
        return None
    return year if 1900 <= year <= 2100 else None


def normalize_timestamp(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc).isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), timezone.utc).isoformat()
    return normalize_text(value) or None


def infer_asset_type(description: str, fa_class: str, fa_subclass: str) -> str | None:
    combined = " ".join(part for part in [description, fa_class, fa_subclass] if part).lower()

    if "yard truck" in combined:
        return None
    if "office" in combined or fa_class in {"OFFICE CO", "OFFICE TRL"}:
        return "office_trailer"
    if (
        "container" in combined
        or "storage container" in combined
        or fa_class in {"STORAGE CO", "STORAG TRL"}
    ):
        return "storage_container"
    if fa_class in {"FLATBD TRL", "REEFER CO", "REEFER TRL"}:
        return "specialty_trailer"
    if any(keyword in combined for keyword in ("flatbed", "reefer", "tanker")):
        return "specialty_trailer"
    if fa_class in {"ROAD TRL", "CARTAG TRL"} or "trailer" in combined:
        return "commercial_box_trailer"
    return None


def infer_status(location_code: str, on_rent: bool, disposed: bool) -> str:
    if disposed or location_code in RETIRED_LOCATION_CODES:
        return "retired"
    if on_rent:
        return "on_rent"
    return "available"


def infer_availability(status: str) -> str:
    if status == "available":
        return "rentable"
    return "unavailable"


def branch_category(code: str) -> str:
    if code in PHYSICAL_BRANCH_CODES:
        return "physical_branch_like"
    if code in SPECIAL_LOCATION_CODES:
        return "special_location"
    if not code:
        return "blank"
    return "unknown_location"


def build_summary(description: str, fa_class: str, fa_subclass: str) -> str:
    if description and description != ".":
        return description
    if fa_subclass and fa_class:
        return f"{fa_class} {fa_subclass}"
    if fa_subclass:
        return fa_subclass
    if fa_class:
        return fa_class
    return "Legacy fixed asset"


def normalize_row(header_index: dict[str, int], row: tuple[Any, ...]) -> dict[str, Any]:
    asset_number = normalize_text(row[header_index["No."]])
    description = normalize_text(row[header_index["Description"]])
    make = normalize_text(row[header_index["Make"]])
    vehicle_year = normalize_year(row[header_index["Vehicle Year"]])
    serial_number = normalize_text(row[header_index["Serial No."]]) or None
    mtrz = normalize_text(row[header_index["MTRZ"]]) or None
    location_code = normalize_text(row[header_index["Location Code"]])
    on_rent = normalize_bool(row[header_index["On Rent"]])
    fa_class = normalize_text(row[header_index["FA Class Code"]])
    fa_subclass = normalize_text(row[header_index["FA Subclass Code"]])
    registration_number = (
        normalize_text(row[header_index["Vehicle Registration No."]]) or None
    )
    acquired = normalize_bool(row[header_index["Acquired"]])
    disposed = normalize_bool(row[header_index["Disposed"]])
    next_service_date = normalize_timestamp(row[header_index["Next Service Date"]])

    asset_type = infer_asset_type(description, fa_class, fa_subclass)
    status = infer_status(location_code, on_rent, disposed)

    supported = bool(asset_number and location_code and asset_type)
    reason_parts = []
    if not asset_number:
        reason_parts.append("missing asset number")
    if not location_code:
        reason_parts.append("missing location code")
    if not asset_type:
        reason_parts.append("unsupported asset type")

    subtype = fa_subclass or fa_class or None
    legacy_payload = {
        "description": description or None,
        "faClassCode": fa_class or None,
        "faSubclassCode": fa_subclass or None,
        "make": make or None,
        "vehicleYear": vehicle_year,
        "vehicleRegistrationNumber": registration_number,
        "mtrz": mtrz,
        "acquired": acquired,
        "disposed": disposed,
        "onRent": on_rent,
        "nextServiceDate": next_service_date,
        "sourceLocationCode": location_code or None,
        "sourceWorkbook": "Fixed Assets",
        "sourceSystem": SOURCE,
    }

    return {
        "assetNumber": asset_number,
        "branchCode": location_code,
        "type": asset_type,
        "subtype": subtype,
        "status": status,
        "availability": infer_availability(status),
        "maintenanceStatus": "clear",
        "serialNumber": serial_number,
        "dimensions": {
            "summary": build_summary(description, fa_class, fa_subclass),
            "legacy": legacy_payload,
        },
        "supported": supported,
        "reason": ", ".join(reason_parts) if reason_parts else None,
        "faClassCode": fa_class,
        "faSubclassCode": fa_subclass,
        "onRent": on_rent,
        "disposed": disposed,
    }


def profile_workbook(source_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if "Fixed Assets" not in workbook.sheetnames:
        raise RuntimeError(
            f'Workbook "{source_path}" does not contain a "Fixed Assets" sheet.'
        )

    worksheet = workbook["Fixed Assets"]
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise RuntimeError(f'Workbook "{source_path}" is empty.')

    headers = [normalize_text(value) for value in header_row]
    if headers != EXPECTED_HEADERS:
        raise RuntimeError(
            "Unexpected workbook headers. "
            f"Expected {EXPECTED_HEADERS}, received {headers}."
        )

    header_index = {header: position for position, header in enumerate(headers)}
    normalized_rows: list[dict[str, Any]] = []
    location_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    unsupported_reason_counts: Counter[str] = Counter()

    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        normalized = normalize_row(header_index, row)
        normalized_rows.append(normalized)
        location_counts[normalized["branchCode"]] += 1
        if normalized["type"]:
            type_counts[normalized["type"]] += 1
        if normalized["supported"]:
            status_counts[normalized["status"]] += 1
        else:
            unsupported_reason_counts[normalized["reason"] or "unsupported"] += 1

    supported_rows = [row for row in normalized_rows if row["supported"]]
    skipped_rows = [row for row in normalized_rows if not row["supported"]]

    location_categories = Counter(branch_category(code) for code in location_counts)
    report = {
        "generatedAt": now_utc().isoformat(),
        "source": str(source_path),
        "workbook": {
            "sheet": "Fixed Assets",
            "rowCount": len(normalized_rows),
            "columnCount": len(headers),
            "headers": headers,
        },
        "coverage": {
            "seedableAssets": len(supported_rows),
            "skippedAssets": len(skipped_rows),
            "supportedFirstClassFields": [
                "assets.assetNumber <- No.",
                "assets.type <- inferred from FA Class Code / FA Subclass Code / Description",
                "assets.subtype <- FA Subclass Code (fallback FA Class Code)",
                "assets.branchId <- Location Code via branch-by-code seed",
                "assets.status <- Disposed / Location Code / On Rent",
                "assets.availability <- derived from status",
                "assets.maintenanceStatus <- conservative default clear",
                "assets.serialNumber <- Serial No.",
            ],
            "preservedInDimensionsLegacyPayload": [
                "Description",
                "Make",
                "Vehicle Year",
                "MTRZ",
                "Vehicle Registration No.",
                "Acquired",
                "Disposed",
                "On Rent",
                "Next Service Date",
                "Location Code",
                "FA Class Code",
                "FA Subclass Code",
            ],
            "notSeededAsFirstClassFields": [
                "customers",
                "customer locations",
                "contracts",
                "contract lines",
                "invoices",
                "invoice lines",
                "GPS device identifiers",
                "Record360 unit identifiers",
                "SkyBitz asset identifiers",
                "maintenance history and work orders",
                "exact manufactured/purchase dates",
            ],
        },
        "branches": {
            "distinctLocationCodes": len(location_counts),
            "distinctSeededBranchCodes": len({row["branchCode"] for row in supported_rows}),
            "locationCategoryCounts": dict(location_categories),
            "topLocationCodes": location_counts.most_common(20),
        },
        "assets": {
            "typeCounts": dict(type_counts),
            "statusCounts": dict(status_counts),
            "unsupportedReasonCounts": dict(unsupported_reason_counts),
        },
        "skippedSamples": [
            {
                "assetNumber": row["assetNumber"],
                "branchCode": row["branchCode"],
                "faClassCode": row["faClassCode"],
                "faSubclassCode": row["faSubclassCode"],
                "reason": row["reason"],
            }
            for row in skipped_rows[:25]
        ],
    }

    return normalized_rows, report


def ensure_output_directory(project_root: Path, output_override: str | None) -> Path:
    if output_override:
        output_dir = Path(output_override)
    else:
        timestamp = now_utc().strftime("%Y-%m-%dT%H-%M-%SZ")
        output_dir = project_root / "artifacts" / "fixed-assets-import" / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def write_report(
    output_dir: Path,
    report: dict[str, Any],
    rows: list[dict[str, Any]],
    write_mode: bool,
    seed_result: dict[str, Any] | None = None,
) -> None:
    supported_rows = [row for row in rows if row["supported"]]
    skipped_rows = [row for row in rows if not row["supported"]]

    summary_payload = {
        "writeMode": write_mode,
        "report": report,
        "seedResult": seed_result,
    }

    (output_dir / "summary.json").write_text(
        json.dumps(summary_payload, indent=2),
        encoding="utf-8",
    )
    (output_dir / "supported-preview.json").write_text(
        json.dumps(supported_rows[:50], indent=2),
        encoding="utf-8",
    )
    (output_dir / "skipped-rows.json").write_text(
        json.dumps(skipped_rows, indent=2),
        encoding="utf-8",
    )


def build_branch_rows(
    supported_rows: list[dict[str, Any]],
    timestamp: datetime,
) -> list[tuple[Any, ...]]:
    branch_rows: list[tuple[Any, ...]] = []
    seen_codes: set[str] = set()

    for row in supported_rows:
        code = row["branchCode"]
        if code in seen_codes:
            continue
        seen_codes.add(code)
        branch_rows.append(
            (
                stable_id("branch_fixed_assets", code),
                code,
                code,
                "America/New_York",
                None,
                None,
                Json(
                    {
                        "source": SOURCE,
                        "placeholder": True,
                        "legacyLocationCode": code,
                        "category": branch_category(code),
                    }
                ),
                timestamp,
                timestamp,
            )
        )

    return branch_rows


def upsert_branches(cursor: Any, supported_rows: list[dict[str, Any]], timestamp: datetime) -> dict[str, str]:
    branch_rows = build_branch_rows(supported_rows, timestamp)
    execute_values(
        cursor,
        """
        INSERT INTO branches (
          id,
          code,
          name,
          timezone,
          phone,
          email,
          address,
          created_at,
          updated_at
        ) VALUES %s
        ON CONFLICT (code) DO UPDATE SET
          name = EXCLUDED.name,
          timezone = EXCLUDED.timezone,
          address = EXCLUDED.address,
          updated_at = EXCLUDED.updated_at
        """,
        branch_rows,
        page_size=250,
    )

    cursor.execute(
        "SELECT id, code FROM branches WHERE code = ANY(%s)",
        ([row["branchCode"] for row in supported_rows],),
    )
    return {code: branch_id for branch_id, code in cursor.fetchall()}


def build_asset_rows(
    supported_rows: list[dict[str, Any]],
    branch_id_by_code: dict[str, str],
    timestamp: datetime,
) -> list[tuple[Any, ...]]:
    asset_rows: list[tuple[Any, ...]] = []

    for row in supported_rows:
        asset_number = row["assetNumber"]
        branch_code = row["branchCode"]
        branch_id = branch_id_by_code[branch_code]
        asset_rows.append(
            (
                stable_id("asset_fixed_assets", asset_number),
                asset_number,
                row["type"],
                row["subtype"],
                Json(row["dimensions"]),
                branch_id,
                row["status"],
                row["availability"],
                row["maintenanceStatus"],
                None,
                row["serialNumber"],
                None,
                [],
                None,
                None,
                None,
                None,
                None,
                None,
                timestamp,
                timestamp,
            )
        )

    return asset_rows


def upsert_assets(
    cursor: Any,
    supported_rows: list[dict[str, Any]],
    branch_id_by_code: dict[str, str],
    timestamp: datetime,
) -> None:
    asset_rows = build_asset_rows(supported_rows, branch_id_by_code, timestamp)
    execute_values(
        cursor,
        """
        INSERT INTO assets (
          id,
          asset_number,
          type,
          subtype,
          dimensions,
          branch_id,
          status,
          availability,
          maintenance_status,
          gps_device_id,
          serial_number,
          manufactured_at,
          features,
          purchase_date,
          yard_zone,
          yard_row,
          yard_slot,
          record360_unit_id,
          skybitz_asset_id,
          created_at,
          updated_at
        ) VALUES %s
        ON CONFLICT (asset_number) DO UPDATE SET
          type = EXCLUDED.type,
          subtype = EXCLUDED.subtype,
          dimensions = EXCLUDED.dimensions,
          branch_id = EXCLUDED.branch_id,
          status = EXCLUDED.status,
          availability = EXCLUDED.availability,
          maintenance_status = EXCLUDED.maintenance_status,
          serial_number = EXCLUDED.serial_number,
          updated_at = EXCLUDED.updated_at
        """,
        asset_rows,
        page_size=500,
    )


def seed_database(
    database_url: str,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    supported_rows = [row for row in rows if row["supported"]]
    timestamp = now_utc()

    connection = psycopg2.connect(database_url)
    try:
        with connection:
            with connection.cursor() as cursor:
                branch_id_by_code = upsert_branches(cursor, supported_rows, timestamp)
                upsert_assets(cursor, supported_rows, branch_id_by_code, timestamp)

                cursor.execute("SELECT COUNT(*) FROM branches")
                branch_count = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM assets")
                asset_count = cursor.fetchone()[0]

                cursor.execute(
                    """
                    SELECT status, COUNT(*)
                    FROM assets
                    GROUP BY status
                    ORDER BY status
                    """
                )
                status_counts = {status: count for status, count in cursor.fetchall()}

        return {
            "seededBranchCodes": len(branch_id_by_code),
            "seededAssets": len(supported_rows),
            "databaseBranchCount": branch_count,
            "databaseAssetCount": asset_count,
            "databaseAssetStatusCounts": status_counts,
        }
    finally:
        connection.close()


def main() -> int:
    args = parse_args()
    project_root = Path.cwd()
    source_path = Path(args.source)
    if not source_path.exists():
        raise RuntimeError(f'Source workbook "{source_path}" does not exist.')

    output_dir = ensure_output_directory(project_root, args.out)
    rows, report = profile_workbook(source_path)

    seed_result = None
    if args.write:
        database_url = require_database_url(project_root)
        seed_result = seed_database(database_url, rows)

    write_report(output_dir, report, rows, args.write, seed_result)

    payload = {
        "outputDirectory": str(output_dir),
        "writeMode": args.write,
        "seedableAssets": report["coverage"]["seedableAssets"],
        "skippedAssets": report["coverage"]["skippedAssets"],
        "distinctSeededBranchCodes": report["branches"]["distinctSeededBranchCodes"],
        "assetStatusCounts": report["assets"]["statusCounts"],
        "seedResult": seed_result,
    }
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover
        print(str(error), file=sys.stderr)
        raise
